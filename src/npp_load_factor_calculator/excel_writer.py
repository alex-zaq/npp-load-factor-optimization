
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image
import pandas as pd

from src.npp_load_factor_calculator.utilites import (
    Converter,
    dict_to_rows,
    get_all_block_repairs_df_by_dict,
    get_file_name_by_scenario,
    get_file_name_with_auto_number,
    get_months_name_by_date_range,
    get_years_by_date_range,
)


class Excel_writer:
    
    def __init__(self, block_grouper, solution_proccesor):
        self.block_grouper = block_grouper
        self.solution_proccesor = solution_proccesor
        self.solver_log = self.solution_proccesor.oemof_model.solver_log
        

    def _write_scenario_options(self, writer, sheet_name):
        scenario_options = self.block_grouper.custom_es.scenario
        
        rows  = dict_to_rows(scenario_options)
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _write_results_data(self, writer, sheet_name):
        
        res = pd.DataFrame()
        
        coeff = Converter.convert(1.0, "мвтч", "млн.квтч")
        el_gen_df = self.block_grouper.get_electricity_profile_all_blocks().resample('M').sum()
        el_gen_df = el_gen_df * coeff * 24
        risk_increase_df = self.block_grouper.get_increase_all_blocks_df().resample('M').mean()
        risk_decrease_df = self.block_grouper.get_decrease_all_blocks_df().resample('M').mean()
        cost_all_blocks_df = self.block_grouper.get_cost_profile_all_blocks(cumulative=False).resample('M').sum()
        cost_all_blocks_df = cost_all_blocks_df.cumsum()
        risks_dict = self.block_grouper.get_risks_profile_by_all_blocks_dict()
        risk_data_dict = {k: v["risk_line_col"] for k, v in risks_dict.items()}
        risk_df = pd.DataFrame(risk_data_dict).resample('M').mean()
        repairs_dict = self.block_grouper.get_repairs_profile_by_all_blocks_dict(part = 0)

        repair_df = get_all_block_repairs_df_by_dict(repairs_dict).resample('M').sum()

        res = pd.concat([el_gen_df, risk_df, risk_increase_df, risk_decrease_df, repair_df, cost_all_blocks_df], axis=1)

        year_col = get_years_by_date_range(res.index)
        months_col = get_months_name_by_date_range(res.index)

        res.insert(0, "year", year_col)
        res.insert(1, "month", months_col)
        
        res.to_excel(writer, sheet_name=sheet_name, index=False)


    def _write_log_info(self, writer, sheet_name):
        solver = self.solution_proccesor.oemof_model.solver
        log_df = pd.DataFrame({f'{solver}_log': self.solver_log.split('\n')})
        log_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    
    
    def add_images(self, images, dpi):
        self.images = images
        self.image_dpi = dpi
        
    

        
        
    def _delete_images(self):
        if hasattr(self, "images"):
            [image.delete_file() for image in self.images]
                   
        
        
    def write(self, folder):
        scen = self.block_grouper.custom_es.scenario
        self.folder = Path(folder)
        if not self.folder.exists():
            self.folder.mkdir(parents=True)
        
        excel_name = get_file_name_by_scenario(scen)
        excel_file = get_file_name_with_auto_number(folder, excel_name, "xlsx")
        path = self.folder / excel_file
        solver_name = self.solution_proccesor.oemof_model.solver

        # Этап 1: Записываем данные
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            self._write_results_data(writer, sheet_name="results")
            self._write_scenario_options(writer, sheet_name="scenario_options")
            self._write_log_info(writer, sheet_name=f"{solver_name}_log")
        
        # Этап 2: Добавляем изображения
        if self.images:
            self._add_images_to_existing_file(path)
        
        self._delete_images()
        print("{}  ({})".format("excel файл создан", excel_file))

    def _add_images_to_existing_file(self, path):
        """Добавляет изображения в существующий файл Excel"""
        from openpyxl import load_workbook
        
        # Сохраняем изображения
        save_img_name_path_pairs = []
        for image in self.images:
            folder_img = str(self.folder.resolve())
            image.save(folder=folder_img, format="jpg", dpi=self.image_dpi)
            save_img_name_path_pairs.append((image.name, image.path))
        
        # Открываем существующий файл
        wb = load_workbook(path)
        
        for img_name, image_path in save_img_name_path_pairs:
            # Создаем новый лист
            ws = wb.create_sheet(img_name)
            # ws = wb.create_sheet("zaq")
            
            # Добавляем изображение
            img = Image(image_path)
            img.anchor = 'A2'
            ws.add_image(img)
        
        # Сохраняем изменения
        wb.save(path)
        wb.close()
    
    